import xml.etree.ElementTree as ET
import re
from openpyxl import load_workbook
from resources import Robot_import

class testcase_generator:
    def __init__(self):
        pass
    
    def parse_excel_to_robot(self,excel_file, robot_file):
        # Load the Excel workbook
        wb = load_workbook(excel_file)
        
        # Assuming the relevant data is in the first sheet
        sheet = wb.active
        
        with open(robot_file, 'w') as rf:
            # Write the headers for the Robot Framework file
            rf.write("*** Settings ***\n")
            # rf.write(f"Library     {'\\nLibrary     '.join(Robot_import.ROBOT_LIB)}\n")
            rf.write("Library     " + "\nLibrary     ".join(Robot_import.ROBOT_LIB) + "\n")
            rf.write("\n*** Test Cases ***")

            test_case_title = None
            documention = []
            keyword = []
            test_setup = []
            test_teardown = []
            for row in sheet.iter_rows(min_row=2, values_only=True):  
                step_type, sl_no, test_case_title, test_step, keywords, args1, args2, args3 = row
                # Handle the start of a new test case
            
                if step_type == "PRECONDITION":
                    rf.write(f"\n{test_case_title}\n")
                    documention.append(f"{test_step}\n")
                    test_setup.append(f"     {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")
                    # rf.write(f"   [Test Setup]    {test_step}")
                    # rf.write(f"   {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")

                # Handle postcondition steps
                elif step_type == "POSTCONDITION":
                    test_teardown.append(f"     {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")    
                    documention.append(f"    ...               {test_step}\n")
                    # rf.write(f"    [Test Teardown]    Setup {test_case_title}\n")
                    # rf.write(f"   {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")

                # Handle exit test case and suite
                elif step_type =="EXIT_TESTCASE":
                    rf.write(f"   [Documentation]    {''.join(documention)}")
                    rf.write(f"   [Setup]{'             '.join(test_setup)}\n")
                    rf.write(f"{''.join(keyword)}\n")
                    rf.write(f"   [Teardown]{'             '.join(test_teardown)}\n")
                    print(f"Exited testcase {sl_no}")
                    documention = []
                    keyword = []
                    test_setup = []
                    test_teardown = []
                elif step_type =="EXIT_TESTSUITE":
                    break
                elif step_type!="TESTCASE_ID":
                    documention.append(f"    ...               {test_step}\n")
                    keyword.append(f"   {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")  
                    # rf.write(f"   {keywords}    {args1 if args1 else ''}    {args2 if args2 else ''}    {args3 if args3 else ''}\n")

        
    def parse_xml(self,xml_path):
        try:
            tree = ET.parse(xml_path)
            root = tree.getroot()
            namespaces = {
                'table': 'urn:oasis:names:tc:opendocument:xmlns:table:1.0',
                'text': 'urn:oasis:names:tc:opendocument:xmlns:text:1.0',
                'office': 'urn:oasis:names:tc:opendocument:xmlns:office:1.0',
            }
            rows = root.findall('.//table:table-row', namespaces)
            testcases_list = []
            for row in rows[1:]: 
                cells = row.findall('table:table-cell', namespaces)
                if len(cells) >= 3:
                    precondition = "\n".join([p.text.strip() for p in cells[1].findall('text:p', namespaces) if p.text])
                    postcondition = "\n".join([p.text.strip() for p in cells[2].findall('text:p', namespaces) if p.text])
                    testcase = "\n".join([p.text.strip() for p in cells[3].findall('text:p', namespaces) if p.text])
                    testcase_dict = {}
                    if precondition:
                        testcase_dict['Preconditions'] = precondition
                    if postcondition:
                        testcase_dict['Postconditions'] = postcondition
                    if testcase:
                        testcase_dict['TestSteps'] = testcase
                    testcases_list.append(testcase_dict)
            return testcases_list
        except Exception as e:
            print(e)
            # raise Exception("Error Occured while parsing")
        
    def _format_steps(self,steps):
        steps = re.sub(r'^\d+\.\s+', '', steps, flags=re.MULTILINE)
        return steps.replace('\n', '\n    ').strip()
        
    def generate_robotframework_testcase(self,test_cases, filename='test_cases.robot'):
        try:
            with open(filename, 'w') as f:
                f.write('*** Settings ***\n')
                f.write("Library     " + "\nLibrary     ".join(Robot_import.ROBOT_LIB) + "\n")
                f.write('\n*** Test Cases ***\n')
                for i, test_case in enumerate(test_cases):
                    test_case_name = f'Test Case {i + 1}'
                    f.write(f'{test_case_name}\n')
                    f.write(f'    [Setup]    Setup {test_case_name}\n')
                    f.write(f'    [Teardown]    Teardown {test_case_name}\n')
                    f.write(f'    {self._format_steps(test_case["TestSteps"])}\n')
                    # f.write(f'    {str(test_case["TestSteps"]).lstrip("1234567890.)")}\n')

                f.write('\n*** Keywords ***\n')
                for i, test_case in enumerate(test_cases):
                    test_case_name = f'Test Case {i + 1}'
                    f.write(f'Setup {test_case_name}\n')
                    preconditions = self._format_steps(test_case['Preconditions'])
                    f.write(f'    {preconditions}\n')

                    f.write(f'Teardown {test_case_name}\n')
                    postconditions = self._format_steps(test_case['Postconditions'])
                    f.write(f'    {postconditions}\n')
            return True
        except Exception as e:
            print(e)
            raise Exception("Error occured while creating testcases")


    # print(parse_excel(r'C:\Users\nfa1cob\Documents\check.xlsx'))
    # a = parse_xml(r'C:\Users\nfa1cob\Documents\check.xml')
    # generate_robotframework_testcase(a)
